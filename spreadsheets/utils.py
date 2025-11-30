"""
Utilities for spreadsheet import/export and operations
"""
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
from io import StringIO, BytesIO


def evaluate_formula(formula, data):
    """
    Simple formula evaluator for spreadsheets
    Supports: =A1, =A1+B1, =SUM(A1:A3), =AVERAGE(A1:A3), etc.
    """
    if not formula.startswith('='):
        return formula
    
    formula = formula[1:].strip()
    
    try:
        # Replace cell references with values
        import re
        cell_pattern = r'[A-Z]+\d+'
        
        def get_cell_value(match):
            cell_ref = match.group(0)
            row, col = _parse_cell_ref(cell_ref)
            if row < len(data) and col < len(data[row]):
                val = data[row][col]
                try:
                    return str(float(val) if val else 0)
                except:
                    return '0'
            return '0'
        
        formula = re.sub(cell_pattern, get_cell_value, formula)
        
        # Handle SUM and AVERAGE functions
        formula = re.sub(r'SUM\(([^)]+)\)', lambda m: str(_sum_range(m.group(1), data)), formula)
        formula = re.sub(r'AVERAGE\(([^)]+)\)', lambda m: str(_average_range(m.group(1), data)), formula)
        
        # Evaluate the expression
        result = eval(formula)
        return str(result)
    except:
        return '#ERROR'


def _parse_cell_ref(cell_ref):
    """Parse cell reference like 'A1' to (row, col)"""
    import re
    match = re.match(r'([A-Z]+)(\d+)', cell_ref)
    if match:
        col_letters = match.group(1)
        row = int(match.group(2)) - 1
        col = 0
        for char in col_letters:
            col = col * 26 + (ord(char) - ord('A') + 1)
        col -= 1
        return row, col
    return 0, 0


def _parse_range(range_ref, data):
    """Parse range like 'A1:A3' and return values"""
    if ':' not in range_ref:
        row, col = _parse_cell_ref(range_ref)
        if row < len(data) and col < len(data[row]):
            return [data[row][col]]
        return [0]
    
    start, end = range_ref.split(':')
    start_row, start_col = _parse_cell_ref(start)
    end_row, end_col = _parse_cell_ref(end)
    
    values = []
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            if r < len(data) and c < len(data[r]):
                val = data[r][c]
                try:
                    values.append(float(val) if val else 0)
                except:
                    pass
    return values


def _sum_range(range_ref, data):
    """Sum values in a range"""
    values = _parse_range(range_ref, data)
    return sum(values)


def _average_range(range_ref, data):
    """Average values in a range"""
    values = _parse_range(range_ref, data)
    return sum(values) / len(values) if values else 0


def xlsx_to_dict(file_path):
    """Convert XLSX file to dictionary format"""
    wb = load_workbook(file_path)
    data = {'sheets': []}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = []
        
        for row in ws.iter_rows(values_only=True):
            sheet_data.append(list(row))
        
        data['sheets'].append({
            'name': sheet_name,
            'data': sheet_data
        })
    
    return data


def dict_to_xlsx(data):
    """Convert dictionary format to XLSX file bytes"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for sheet_info in data.get('sheets', []):
        ws = wb.create_sheet(sheet_info['name'])
        sheet_data = sheet_info.get('data', [])
        
        for row_idx, row_data in enumerate(sheet_data, 1):
            for col_idx, cell_value in enumerate(row_data or [], 1):
                ws.cell(row_idx, col_idx, cell_value)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def dict_to_csv(data, sheet_index=0):
    """Convert dictionary format to CSV format"""
    sheets = data.get('sheets', [])
    if sheet_index >= len(sheets):
        sheet_index = 0
    
    sheet_data = sheets[sheet_index].get('data', [])
    
    output = StringIO()
    writer = csv.writer(output)
    
    for row in sheet_data:
        writer.writerow(row or [])
    
    return output.getvalue()


def initialize_spreadsheet():
    """Create initial empty spreadsheet structure"""
    return {
        'sheets': [{
            'name': 'Sheet1',
            'data': [[''] * 10 for _ in range(20)]  # 20 rows x 10 columns
        }]
    }
